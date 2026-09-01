from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


OUTPUT = Path("subject_distribution_ocr.xlsx")
SOURCE = Path(r"C:\Users\Lenovo\Desktop\WhatsApp Image 2026-09-01 at 12.05.04 AM.jpeg")
ENHANCED = Path(".runtime/subject-distribution-enhanced.png")

classes = ["SE", "ES", "SV", "LS", "ES2", "11B", "11A", "ES1", "10B", "10A", "EB9", "9B", "9A", "EB8", "8B", "8A", "EB7", "7B", "7A"]

# Values are a careful transcription of the photographed grid. Lower social-
# science rows and several overwritten cells are explicitly low confidence.
matrix = {
    "Mathematics": [5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 6, 6, 6, 6, 6, 5, 6, 6],
    "Physics": [1, 1, 6, 6, 5, 4, 4, 4, 4, 4, 3, 2, 2, 2, 2, 2, 3, 2, 2],
    "Arabic": [3, 3, 3, 2, 3, 3, 3, 3, 3, 3, 7, 6, 6, 7, 7, 7, 6, 6, 6],
    "English": [5, 5, 6, 6, 6, 6, 6, 5, 5, 5, 7, 7, 7, 7, 7, 7, 7, 7, 7],
    "Chemistry": [1, 1, 4, 4, 3, 3, 3, 4, 2, 2, 2, 3, 3, 2, 2, 2, 2, 2, 2],
    "History": [1] * 19,
    "Geography": [1] * 19,
    "Civics / Education": [1] * 19,
    "Biology": [1, 1, 6, 6, 3, 3, 3, 2, 2, 2, 3, 2, 2, 2, 2, 2, 2, 2, 2],
    "Economics": [6, 6, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
    "Sociology": [3, 3, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
    "Philosophy": [3, 3, 2, 2, 1, 1, 1, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
}

low_confidence = {
    (subject, cls)
    for subject in ("Economics", "Sociology", "Philosophy")
    for cls in classes
}
low_confidence.update({
    ("Mathematics", "EB8"), ("Mathematics", "EB7"),
    ("Physics", "ES2"), ("Physics", "11A"), ("Physics", "11B"),
    ("Arabic", "EB9"), ("English", "ES2"), ("Chemistry", "ES1"),
})

wb = Workbook()
grid = wb.active
grid.title = "Subject Matrix"
grid.append(["Subject / Class", *classes, "Row total"])
for subject, values in matrix.items():
    row = grid.max_row + 1
    grid.append([subject, *values, f"=SUM(B{row}:T{row})"])

normalized = wb.create_sheet("Normalized Data")
normalized.append(["Class", "Subject", "Sessions per week", "Confidence", "Needs verification"])
for subject, values in matrix.items():
    for cls, value in zip(classes, values, strict=True):
        low = (subject, cls) in low_confidence
        normalized.append([cls, subject, value, "Low" if low else "Medium", "Yes" if low else "No"])

totals = wb.create_sheet("Class Totals")
totals.append(["Class", "Transcribed subject total", "Important note"])
for cls_index, cls in enumerate(classes, start=2):
    letter = get_column_letter(cls_index)
    totals.append([cls, f"=SUM('Subject Matrix'!{letter}2:{letter}{len(matrix)+1})", "Only subjects visible in the photographed matrix are counted."])

notes = wb.create_sheet("OCR Notes")
notes.append(["Item", "Details"])
notes.append(["Source", str(SOURCE)])
notes.append(["Method", "Rotated and contrast-enhanced; Tesseract OCR used for Latin codes; handwritten Arabic numerals manually reconciled."])
notes.append(["Warning", "The photo is faint and folded. Orange cells are uncertain and must be checked against the paper before database import."])
notes.append(["Zeros", "A zero means the photographed grid appeared blank/not applicable; low-confidence zeros may represent unreadable writing."])
notes.append(["Scope", "The workbook contains the visible core matrix. Notes written below the grid were not treated as class-subject allocations."])

source_sheet = wb.create_sheet("Enhanced Source")
source_sheet["A1"] = "Enhanced working image used for transcription"
source_sheet["A1"].font = Font(bold=True)
if ENHANCED.exists():
    image = ExcelImage(ENHANCED)
    image.width = 1150
    image.height = 480
    source_sheet.add_image(image, "A3")

header_fill = PatternFill("solid", fgColor="1F4E78")
low_fill = PatternFill("solid", fgColor="F4B183")
for sheet in (grid, normalized, totals, notes):
    sheet.freeze_panes = "B2" if sheet is grid else "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

for row_idx, subject in enumerate(matrix, start=2):
    for col_idx, cls in enumerate(classes, start=2):
        if (subject, cls) in low_confidence:
            grid.cell(row_idx, col_idx).fill = low_fill

for row in normalized.iter_rows(min_row=2):
    if row[3].value == "Low":
        for cell in row:
            cell.fill = low_fill

grid.column_dimensions["A"].width = 24
for col in range(2, 21):
    grid.column_dimensions[get_column_letter(col)].width = 8
grid.column_dimensions["U"].width = 12
normalized.column_dimensions["A"].width = 12
normalized.column_dimensions["B"].width = 24
normalized.column_dimensions["C"].width = 20
normalized.column_dimensions["D"].width = 14
normalized.column_dimensions["E"].width = 20
totals.column_dimensions["A"].width = 12
totals.column_dimensions["B"].width = 25
totals.column_dimensions["C"].width = 65
notes.column_dimensions["A"].width = 18
notes.column_dimensions["B"].width = 110

wb.save(OUTPUT)
print(OUTPUT.resolve())
