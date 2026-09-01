from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


OUTPUT = Path("teacher_assignments_ocr.xlsx")
SOURCE = r"C:\Users\Lenovo\Desktop\WhatsApp Image 2026-08-31 at 11.08.53 PM.jpeg"

teachers = [
    ("T01", "نادين شرف", "Low", 29, "Name and overwritten cells need confirmation."),
    ("T02", "رول جوزف عري", "Low", 23, "Name and several covered cells need confirmation."),
    ("T03", "سيدا غنمي", "Low", 24, "Name uncertain; class/hour row is comparatively clear."),
    ("T04", "ليلى السيد", "Medium", 28, "French row; name appears readable."),
    ("T05", "جنى بلال", "Medium", 10, "Mathematics row."),
    ("T06", "علي شحادة", "Low", 10, "Surname uncertain."),
    ("T07", "حنان خليفة", "Medium", 24, "Chemistry row; one digit interpreted from stated total."),
    ("T08", "الاسم غير واضح", "Low", 6, "Teacher name is not legible."),
    ("T09", "فداء (الكنية غير واضحة)", "Low", 11, "Surname and one class code need confirmation."),
    ("T10", "محمد عثمان", "Medium", 14, "Physics row."),
    ("T11", "محمد (الكنية غير واضحة)", "Low", 18, "Biology row; surname unclear."),
    ("T12", "كريستين (الكنية غير واضحة)", "Low", 21, "Social studies row; surname unclear."),
    ("T13", "خالد (الكنية غير واضحة)", "Low", 22, "History row; surname unclear."),
    ("T14", "محمد (الاسم الكامل غير واضح)", "Low", 12, "Bottom row is partly outside/obscured in the photograph."),
]

assignments = [
    ("T01", "Physics", "7A", 2, "High"), ("T01", "Physics", "EB7", 3, "Medium"),
    ("T01", "Mathematics", "EB7", 5, "Medium"), ("T01", "Mathematics", "EB8", 4, "High"),
    ("T01", "Physics", "EB8", 2, "Medium"), ("T01", "Mathematics", "EB9", 5, "Low"),
    ("T01", "Physics", "EB9", 3, "Medium"), ("T01", "Chemistry", "EB9", 2, "Low"),
    ("T01", "Unresolved", "Covered/unclear", 3, "Low"),
    ("T02", "Unknown", "9B", 6, "Medium"), ("T02", "Unknown", "EB9", 4, "Low"),
    ("T02", "Unknown", "SV", 3, "Medium"), ("T02", "Unknown", "LS", 2, "Medium"),
    ("T02", "Unknown", "SE", 3, "Medium"), ("T02", "Unresolved", "Covered/unclear", 5, "Low"),
    ("T03", "English", "10A", 5, "High"), ("T03", "English", "10B", 5, "High"),
    ("T03", "English", "11A", 3, "High"), ("T03", "English", "11B", 3, "High"),
    ("T03", "English", "ES", 5, "Medium"), ("T03", "English", "LS", 3, "Medium"),
    ("T04", "French", "EB9", 7, "High"), ("T04", "French", "EB8", 6, "High"),
    ("T04", "French", "ES2", 3, "Medium"), ("T04", "French", "ES1", 5, "Medium"),
    ("T04", "French", "SE", 5, "Medium"), ("T04", "French", "SV", 2, "Low"),
    ("T05", "Mathematics", "ES1", 5, "High"), ("T05", "Mathematics", "ES2", 5, "High"),
    ("T06", "Mathematics", "SV", 5, "High"), ("T06", "Mathematics", "SE", 5, "High"),
    ("T07", "Chemistry", "LS", 4, "High"), ("T07", "Chemistry", "ES", 2, "Medium"),
    ("T07", "Chemistry", "10A", 3, "High"), ("T07", "Chemistry", "10B", 3, "Medium"),
    ("T07", "Chemistry", "11A", 3, "High"), ("T07", "Chemistry", "11B", 3, "High"),
    ("T07", "Chemistry", "9A", 3, "High"), ("T07", "Chemistry", "9B", 3, "High"),
    ("T08", "Social studies", "10A", 1, "High"), ("T08", "Social studies", "10B", 1, "High"),
    ("T08", "Social studies", "ES1", 1, "Medium"), ("T08", "Social studies", "11A", 1, "High"),
    ("T08", "Social studies", "11B", 1, "High"), ("T08", "Social studies", "ES2", 1, "Medium"),
    ("T09", "Philosophy", "LS", 2, "Medium"), ("T09", "Philosophy", "SV", 2, "Medium"),
    ("T09", "Philosophy", "ES", 2, "Medium"), ("T09", "Philosophy", "SE", 2, "Medium"),
    ("T09", "Philosophy", "11A", 1, "High"), ("T09", "Philosophy", "11B", 1, "High"),
    ("T09", "Philosophy", "ES2", 1, "Low"),
    ("T10", "Physics", "L (code unclear)", 6, "Low"), ("T10", "Physics", "10A", 4, "High"),
    ("T10", "Physics", "10B", 4, "High"),
    ("T11", "Biology", "LS", 6, "High"), ("T11", "Physics", "ES", 1, "Medium"),
    ("T11", "Biology", "ES", 1, "Medium"), ("T11", "Biology", "10A", 2, "High"),
    ("T11", "Biology", "10B", 2, "High"), ("T11", "Biology", "11A", 3, "High"),
    ("T11", "Biology", "11B", 3, "High"),
    ("T12", "Social studies", "ES", 6, "Medium"), ("T12", "Social studies", "SE", 6, "Medium"),
    ("T12", "Social studies", "10A", 1, "High"), ("T12", "Social studies", "10B", 1, "High"),
    ("T12", "Social studies", "ES1", 1, "Medium"), ("T12", "Social studies", "11A", 2, "High"),
    ("T12", "Social studies", "11B", 2, "High"), ("T12", "Social studies", "ES2", 2, "Medium"),
    ("T13", "History", "LS", 1, "Medium"), ("T13", "History", "SV", 1, "Medium"),
    ("T13", "History", "ES", 1, "Medium"), ("T13", "History", "SE", 1, "Medium"),
    ("T13", "History", "10A", 3, "High"), ("T13", "History", "10B", 3, "High"),
    ("T13", "History", "ES1", 3, "Medium"), ("T13", "History", "11A", 3, "High"),
    ("T13", "History", "11B", 3, "High"), ("T13", "History", "ES2", 3, "Medium"),
    ("T14", "Unresolved", "Bottom row obscured", 12, "Low"),
]

wb = Workbook()
ws = wb.active
ws.title = "Assignments"
headers = ["Teacher ID", "Teacher name (OCR/manual)", "Subject", "Class", "Hours", "Reading confidence"]
ws.append(headers)
teacher_names = {row[0]: row[1] for row in teachers}
for teacher_id, subject, class_code, hours, confidence in assignments:
    ws.append([teacher_id, teacher_names[teacher_id], subject, class_code, hours, confidence])

summary = wb.create_sheet("Teacher Summary")
summary.append(["Teacher ID", "Teacher name (OCR/manual)", "Name confidence", "Subjects", "Extracted hours", "Stated overall count", "Difference", "Notes"])
for row_index, (teacher_id, name, confidence, stated_total, notes) in enumerate(teachers, start=2):
    subjects = sorted({a[1] for a in assignments if a[0] == teacher_id and a[1] != "Unresolved"})
    summary.append([teacher_id, name, confidence, ", ".join(subjects), f'=SUMIF(Assignments!A:A,A{row_index},Assignments!E:E)', stated_total, f'=E{row_index}-F{row_index}', notes])

notes = wb.create_sheet("OCR Notes")
notes_rows = [
    ("Source image", SOURCE),
    ("Method", "Image rotated/enlarged; Tesseract OCR (English) used for Latin codes; Arabic handwriting manually reconciled."),
    ("Important", "Low-confidence names/cells are explicitly marked. Do not import those rows into production without checking the paper/original author."),
    ("Class codes", "Codes are retained as written (for example EB7, EB8, EB9, ES1, ES2, SE, SV, LS). Their institutional meaning was not inferred."),
    ("Totals", "Unresolved placeholder rows preserve the stated overall total where covered or unreadable assignments prevented a complete breakdown."),
]
for key, value in notes_rows:
    notes.append([key, value])

header_fill = PatternFill("solid", fgColor="1F4E78")
low_fill = PatternFill("solid", fgColor="FCE4D6")
for sheet in (ws, summary):
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        if any(str(cell.value) == "Low" for cell in row):
            for cell in row:
                cell.fill = low_fill

for sheet in wb.worksheets:
    for column in range(1, sheet.max_column + 1):
        max_length = max(len(str(sheet.cell(row, column).value or "")) for row in range(1, sheet.max_row + 1))
        sheet.column_dimensions[get_column_letter(column)].width = min(max(max_length + 2, 12), 48)
    sheet.sheet_view.rightToLeft = False

notes.column_dimensions["A"].width = 20
notes.column_dimensions["B"].width = 110
notes["A1"].font = Font(bold=True)
wb.save(OUTPUT)
print(OUTPUT.resolve())
