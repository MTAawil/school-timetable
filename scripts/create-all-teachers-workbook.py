from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo


OUTPUT = Path("all_school_teachers.xlsx")

# page, teacher, weekly hours in file, assigned classes, final hours, matches
rows = [
    (1, "اميرة قمر", 29, "Grade 10 B، Grade 8 A، Grade 8 B، Grade 9 A، Grade 9 B", 29, "نعم"),
    (1, "رؤى الحاج حسن", 18, "Grade 11 A، Grade 11 B، Grade 7 B، Grade 8 A، Grade 8 B، Grade 9", 18, "نعم"),
    (1, "محمد عبدو", 12, "A، Grade 9 B", 12, "نعم"),
    (1, "ريم حمادة", 27, "GRADE 12 ES، LS، SE، SV", 27, "نعم"),
    (1, "بتول كريب", 24, "Grade 2 A، Grade 2 B، Grade 3 A", 24, "نعم"),
    (1, "غنى بيضون", 25, "Grade 3 A، Grade 3 B، Grade 4 A، Grade 4 C، Grade 5 A", 25, "نعم"),
    (1, "ريتا حاوي", 25, "Grade 3 B، Grade 4 A، Grade 4 C", 25, "نعم"),
    (1, "نور منقري", 25, "Grade 4 A، Grade 4 B، Grade 5 C، Grade 5 B", 25, "نعم"),
    (1, "اماني زين الدين", 25, "Grade 4 A، Grade 4 B، Grade 5 A", 25, "نعم"),
    (1, "نسرين رضا", 24, "Grade 4 B، Grade 4 C، Grade 5 C، Grade 5 B، Grade 6 C، Grade 6 B", 24, "نعم"),
    (1, "الاء فخر الدين", 26, "Grade 6 A، Grade 7 A، Grade 7 B، Grade 8 A، Grade 8 B، Grade 9 A", 26, "نعم"),
    (1, "عزيزة غص", 27, "Grade 6 A، Grade 6 C، Grade 6 B، Grade 7 A، Grade 7 B", 27, "نعم"),
    (1, "سوزان عيسى", 26, "Grade 8 A، Grade 8 B، Grade 9 A، Grade 9 B", 26, "نعم"),
    (1, "علي نمر", 10, "SE، SV", 10, "نعم"),
    (1, "علا حمادة", 8, "EB 5، EB 6، Grade 5 A، Grade 5 B، Grade 5 C، Grade 6 A، Grade 6 B، Grade 6 C", 8, "نعم"),

    (2, "هدى زين الدين", 26, "EB 3، EB1، EB2", 26, "نعم"),
    (2, "اميرة حمية", 22, "EB 4، EB 5، EB 6، EB1، EB2، Grade 1 A، Grade 1 B، Grade 2 A، Grade 2 B، Grade 3 A، Grade 3 B، Grade 4 A، Grade 4 B، Grade 4 C، Grade 5 A، Grade 5 C، Grade 5 B", 22, "نعم"),
    (2, "هنادي نون", 25, "EB 4، EB2، Grade 3 A، Grade 3 B", 25, "نعم"),
    (2, "عدنان زرق", 18, "EB 4، EB1، EB2، Grade 1 A، Grade 1 B، Grade 2 A، Grade 2 B، Grade 3 A، Grade 3 B، Grade 4 A، Grade 4 B، Grade 4 C", 18, "نعم"),
    (2, "فاطمة يحقوفي", 20, "EB 4، EB 5، EB 6، EB1، EB2، Grade 1 A، Grade 1 B، Grade 2 A، Grade 2 B، Grade 3 A، Grade 3 B، Grade 4 A، Grade 4 B، Grade 4 C، Grade 5 A، Grade 5 C، Grade 5 B، Grade 6 A، Grade 6 C، Grade 6 B", 20, "نعم"),
    (2, "هديل قنبر", 26, "EB 4، EB 5، EB 6، Grade 4 A، Grade 4 B، Grade 4 C، Grade 5 A، Grade 5 C، Grade 5 B، Grade 6 A، Grade 6 C، Grade 6 B", 26, "نعم"),
    (2, "هالة حلوان", 24, "EB 5، EB1، Grade 1 A، Grade 1 B، Grade 2 A، Grade 2 B", 24, "نعم"),
    (2, "زهراء الشيخ", 28, "EB 5، EB 6، Grade 5 C، Grade 5 B", 28, "نعم"),
    (2, "آية ناصر", 22, "EB 6، Grade 5 A، Grade 6 A، Grade 6 C", 22, "نعم"),
    (2, "ليلى شمس", 21, "EB 8، Grade 8 A، Grade 8 B", 21, "نعم"),
    (2, "وصال يوسف", 23, "EB 9، GRADE 12 ES، Grade 9 B، LS، SE، SV", 23, "نعم"),
    (2, "كاتيا ابراهيم", 7, "EB1، Grade 4 A، Grade 4 B", 7, "نعم"),
    (2, "حسناء الحسين", 24, "EB2، Grade 2 A، Grade 2 B", 24, "نعم"),
    (2, "منى وهبي مصلحة", 11, "ES2، Grade 11 A، Grade 11 B، GRADE 12 ES، LS، SE، SV", 11, "نعم"),
    (2, "زهراء طالب", 25, "Grade 1 A، Grade 1 B، Grade 2 A، Grade 2 B، Grade 4 B", 25, "نعم"),
    (2, "اسراء فارس", 26, "Grade 1 A، Grade 1 B، Grade 2 A، Grade 2 B، Grade 3 A، Grade 3 B، Grade 5 A، Grade 5 C، Grade 5 B", 26, "نعم"),
    (2, "حنان بليطة", 23, "Grade 10 A، Grade 10 B، Grade 11 A، Grade 11 B، GRADE 12 ES، Grade 9 A، Grade 9 B، LS", 23, "نعم"),
    (2, "ريان جلول", 25, "Grade 10 A، Grade 11 A، Grade 11 B، GRADE 12 ES، LS", 25, "نعم"),
    (2, "محمد جعفرة", 18, "Grade 10 A، Grade 10 B، Grade 11 A، Grade 11 B، GRADE 12 ES، LS", 18, "نعم"),
    (2, "محمد عساف", 14, "Grade 10 A، Grade 10 B، LS", 14, "نعم"),
    (2, "ميساء غصن", 23, "Grade 10 A، Grade 10 B، Grade 11 A، Grade 11 B، GRADE 12 ES، LS", 23, "نعم"),

    (3, "محمد جمال الدين", 6, "EB 7، EB 8، Grade 7 A، Grade 7 B، Grade 8 A، Grade 8 B", 6, "نعم"),
    (3, "حوراء ابو حمدان", 25, "EB 7، Grade 6 C، Grade 7 A، Grade 7 B", 25, "نعم"),
    (3, "نادين شرف", 29, "EB 7، EB 8، EB 9، Grade 7 A، Grade 8 A، Grade 8 B", 29, "نعم"),
    (3, "رانيا ابراهيم", 29, "EB 7، EB 8، EB 9، ES 1، Grade 7 A", 29, "نعم"),
    (3, "رباب السيد", 10, "EB 7، EB 4، EB 5، EB 6", 10, "نعم"),
    (3, "عبير نون", 27, "EB 7، EB 8، EB 9، Grade 7 A، Grade 7 B، Grade 8 A، Grade 8 B", 27, "نعم"),
    (3, "حسن ناجح", 13, "Grade 9 A، Grade 9 B، EB 7، ES 2، EB 8، EB 9، SE، SV", 13, "نعم"),
    (3, "زينة علي", 20, "EB 7، ES 2، EB 5، EB 6، EB 8، ES 1، Grade 10 A، Grade 10 B، Grade 11 A، Grade 11 B، Grade 5 A، Grade 5 B، Grade 6 A، Grade 6 C، Grade 6 B، Grade 7 A، Grade 7 B، Grade 8 A، Grade 8 B", 20, "نعم"),
    (3, "زينب سبور", 24, "EB1، Grade 1 A، GRADE 1 B", 24, "نعم"),
    (3, "احمد الحركة", 6, "ES 2، ES 1، Grade 10 A، Grade 10 B، Grade 11 A، Grade 11 B", 6, "نعم"),
    (3, "محمد المصري", 22, "ES 2، ES 1، Grade 10 A، Grade 10 B، Grade 11 A، Grade 11 B، GRADE 12 ES، LS، SE، SV، Grade 9 A، Grade 9 B", 24, "نعم"),
    (3, "زهراء مصطفى", 27, "ES 2، ES 1، Grade 10 A، Grade 10 B، Grade 11 A، Grade 11 B، Grade 9 A", 27, "نعم"),
    (3, "سحر رفيق انصار", 21, "ES 2، ES 1، Grade 10 A، Grade 10 B، Grade 11 A، Grade 11 B، GRADE 12 ES، SE", 21, "نعم"),
    (3, "صبحي حمية", 16, "ES 2، ES 1، SE، SV", 16, "نعم"),
    (3, "حسن بزيك", 10, "ES 2، ES 1", 10, "نعم"),
    (3, "ليلى السيد", 28, "ES 2، EB 8، EB 9، ES 1، SE، SV", 28, "نعم"),
    (3, "تاجي هاشم", 9, "ES 2، SV", 9, "نعم"),
    (3, "ريما عيسى النجار", 6, "ES 2، ES 1، Grade 10 A، Grade 10 B، Grade 11 A، Grade 11 B", 6, "نعم"),
    (3, "لينا خليل", 25, "EB 3، EB 5، EB 6، EB2", 25, "نعم"),
    (3, "نور سهيل", 26, "EB 3، EB 4، EB 5، EB 6، EB1، EB2، Grade 4 C، Grade 5 A، Grade 5 C، Grade 5 B، Grade 6 A، Grade 6 C، Grade 6 B", 26, "نعم"),
    (3, "زينب محسن", 25, "EB 3، EB 4، EB1، EB2، Grade 6 B، Grade 7 A، Grade 7 B", 25, "نعم"),
]

wb = Workbook()
ws = wb.active
ws.title = "جميع المعلمين"
ws.sheet_view.rightToLeft = True
headers = ["الأستاذ", "الساعات الأسبوعية في الملف", "الصفوف المستلمة", "العدد النهائي للحصص", "مطابقة؟"]
ws.append(headers)
for _, teacher, weekly, classes, final, matches in rows:
    ws.append([teacher, weekly, classes, final, matches])

table = Table(displayName="AllTeachers", ref=f"A1:E{ws.max_row}")
table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
ws.add_table(table)
ws.freeze_panes = "A2"
ws.auto_filter.ref = ws.dimensions
ws.column_dimensions["A"].width = 24
ws.column_dimensions["B"].width = 25
ws.column_dimensions["C"].width = 105
ws.column_dimensions["D"].width = 24
ws.column_dimensions["E"].width = 13
for cell in ws[1]:
    cell.fill = PatternFill("solid", fgColor="1F4E78")
    cell.font = Font(color="FFFFFF", bold=True)
for row in ws.iter_rows():
    for cell in row:
        cell.alignment = Alignment(horizontal="right" if cell.column in (1, 3, 5) else "center", vertical="center", wrap_text=True)
for row_number in range(2, ws.max_row + 1):
    ws.row_dimensions[row_number].height = 36
    if ws.cell(row_number, 2).value != ws.cell(row_number, 4).value:
        for cell in ws[row_number]:
            cell.fill = PatternFill("solid", fgColor="FFF2CC")

audit = wb.create_sheet("مصدر ومراجعة")
audit.sheet_view.rightToLeft = True
audit.append(["رقم الصفحة", "رقم الصف في Excel", "الأستاذ", "ملاحظة"])
for index, (page, teacher, *_rest) in enumerate(rows, start=2):
    note = "العدد النهائي مختلف عن عدد الملف" if ws.cell(index, 2).value != ws.cell(index, 4).value else ""
    audit.append([page, index, teacher, note])
audit.freeze_panes = "A2"
audit.column_dimensions["A"].width = 14
audit.column_dimensions["B"].width = 18
audit.column_dimensions["C"].width = 25
audit.column_dimensions["D"].width = 45

for page in (1, 2, 3):
    sheet = wb.create_sheet(f"المصدر {page}")
    sheet.sheet_view.rightToLeft = True
    sheet["A1"] = f"صورة المصدر رقم {page}"
    sheet["A1"].font = Font(bold=True)
    image_path = Path(f".runtime/teachers-clean-{page}.png")
    if image_path.exists():
        image = ExcelImage(image_path)
        image.width = 1000
        image.height = int(image.height * (1000 / image.width))
        sheet.add_image(image, "A3")

wb.save(OUTPUT)
print(OUTPUT.resolve())
print(f"rows={len(rows)}")
